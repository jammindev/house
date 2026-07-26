"""Generic XLSX statement importer.

Same mapping contract as ``generic_csv`` — only the way rows are produced
differs. ``openpyxl`` gives us real ``date``/``datetime`` and numeric cells,
which ``parsing.parse_date`` / ``parse_amount`` accept directly, so a bank that
exports Excel avoids the string-mangling guesswork entirely.
"""
from __future__ import annotations

import io

from .base import BaseStatementImporter, ImporterError, ImporterFormatError, NormalizedTransaction
from .mapping import build_mapping, find_header_row, rows_to_transactions
from .registry import register

XLSX_MAGIC = b"PK\x03\x04"
SHEET_SCAN_LIMIT = 30


def _load_rows(raw: bytes, options: dict | None) -> list[list]:
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a varied zoo of zip/format errors
        raise ImporterFormatError(f"unreadable XLSX file: {exc}")

    try:
        sheet_name = str((options or {}).get("sheet") or "").strip()
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ImporterError(
                    f"sheet {sheet_name!r} not found (sheets: {', '.join(workbook.sheetnames)})"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


class GenericStatementXlsxImporter(BaseStatementImporter):
    key = "generic_xlsx"
    label = "Excel générique (mapping manuel)"

    def detect(self, raw: bytes) -> bool:
        # An XLSX can only be read one way, so claiming the ZIP magic is safe —
        # unlike CSV, where the delimiter and columns are pure guesswork. The
        # column mapping is still required to parse it.
        return raw[:4] == XLSX_MAGIC

    def parse(self, raw: bytes, *, options: dict | None = None) -> list[NormalizedTransaction]:
        mapping = build_mapping(options)
        rows = _load_rows(raw, options)
        if not rows:
            raise ImporterError("empty file")

        skip_rows = (options or {}).get("skip_rows")
        header_index = find_header_row(
            rows, mapping, skip_rows=int(skip_rows) if skip_rows is not None else None
        )
        return rows_to_transactions(
            rows[header_index],
            rows[header_index + 1 :],
            mapping,
            first_line_no=header_index + 2,
        )

    def columns(self, raw: bytes, *, options: dict | None = None) -> list[str]:
        rows = _load_rows(raw, options)
        if not rows:
            return []
        candidate = max(rows[:SHEET_SCAN_LIMIT], key=len)
        return [str(cell or "").strip() for cell in candidate if str(cell or "").strip()]

    def sample_lines(self, raw: bytes) -> list[str]:
        rows = _load_rows(raw, None)
        return [
            " | ".join("" if cell is None else str(cell) for cell in row) for row in rows[:10]
        ]


register(GenericStatementXlsxImporter())
