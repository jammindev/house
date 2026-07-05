"""Enedis XLSX importer — structure mirrored from a real household export."""
import datetime as dt
import io
from zoneinfo import ZoneInfo

import pytest

from electricity import services
from electricity.importers import detect_importer
from electricity.importers.base import ImporterError
from electricity.importers.enedis_xlsx import EnedisXlsxImporter
from electricity.models import ConsumptionRecord, ImportStatus
from electricity.tests.factories import ElectricityMeterFactory, HouseholdFactory, UserFactory

PARIS = ZoneInfo("Europe/Paris")


def build_xlsx(rows, *, sheet_title="Export Consommation Quotidienne", header=("Date", "Valeur (en kWh)")):
    """Build an in-memory workbook shaped like the real portal export: a cover
    sheet, then the data sheet with metadata lines above the header row."""
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active.title = "Page d'accueil"
    workbook.active.cell(row=6, column=2, value="Les données de ce fichier sont…")
    sheet = workbook.create_sheet(sheet_title)
    sheet.cell(row=8, column=2, value="Point Référence Mesure (PRM) :")
    sheet.cell(row=8, column=3, value="09000000000000")
    sheet.cell(row=14, column=2, value=header[0])
    sheet.cell(row=14, column=3, value=header[1])
    for i, (day, value) in enumerate(rows, start=15):
        sheet.cell(row=i, column=2, value=day)
        sheet.cell(row=i, column=3, value=value)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


DAILY_ROWS = [("05/06/2026", "15.064"), ("06/06/2026", "12.485"), ("07/06/2026", "12.406")]


class TestParse:
    def test_daily_export_parses_to_local_midnight_day_points(self):
        importer = EnedisXlsxImporter()
        points = importer.parse(build_xlsx(DAILY_ROWS), tz=PARIS)
        assert len(points) == 3
        first = points[0]
        assert first.ts_start == dt.datetime(2026, 6, 5, 0, 0, tzinfo=PARIS)
        assert first.interval_minutes == 1440
        assert first.energy_wh == 15064
        assert first.register == "base"
        assert sum(p.energy_wh for p in points) == 15064 + 12485 + 12406

    def test_detected_automatically_and_before_csv(self):
        raw = build_xlsx(DAILY_ROWS)
        importer = detect_importer(raw)
        assert importer is not None and importer.key == "enedis_xlsx"

    def test_gap_rows_and_na_skipped(self):
        rows = [("05/06/2026", "1.0"), ("06/06/2026", None), ("07/06/2026", "NA"), ("08/06/2026", "2.0")]
        points = EnedisXlsxImporter().parse(build_xlsx(rows), tz=PARIS)
        assert [p.energy_wh for p in points] == [1000, 2000]

    def test_numeric_and_datetime_cells_supported(self):
        rows = [(dt.datetime(2026, 6, 5), 15.064), (dt.datetime(2026, 6, 6), 12.485)]
        points = EnedisXlsxImporter().parse(build_xlsx(rows), tz=PARIS)
        assert points[0].ts_start == dt.datetime(2026, 6, 5, 0, 0, tzinfo=PARIS)
        assert points[0].energy_wh == 15064

    def test_unreadable_value_fails_with_row_number(self):
        rows = [("05/06/2026", "1.0"), ("06/06/2026", "abc")]
        with pytest.raises(ImporterError, match="row 16"):
            EnedisXlsxImporter().parse(build_xlsx(rows), tz=PARIS)

    def test_missing_header_is_format_error(self):
        raw = build_xlsx([], header=("Foo", "Bar"))
        with pytest.raises(ImporterError, match="Date / Valeur"):
            EnedisXlsxImporter().parse(raw, tz=PARIS)
        assert EnedisXlsxImporter().detect(raw) is False

    def test_non_xlsx_bytes_not_detected(self):
        assert EnedisXlsxImporter().detect(b"Horodate;Valeur\n2026-06-01T00:30:00+02:00;420\n") is False

    def test_sample_lines_render_rows(self):
        lines = EnedisXlsxImporter().sample_lines(build_xlsx(DAILY_ROWS))
        assert lines[0] == "05/06/2026 ; 15.064"


@pytest.mark.django_db
class TestServiceIntegration:
    def test_import_via_service_is_idempotent(self):
        household, user = HouseholdFactory(), UserFactory()
        meter = ElectricityMeterFactory(household=household, timezone="Europe/Paris")
        upload = io.BytesIO(build_xlsx(DAILY_ROWS))
        upload.name = "export.xlsx"
        imported = services.import_consumption_file(household, user, meter=meter, uploaded_file=upload)
        assert imported.status == ImportStatus.COMPLETED
        assert imported.provider == "enedis_xlsx"
        assert imported.created_count == 3
        assert ConsumptionRecord.objects.filter(meter=meter).count() == 3

        upload2 = io.BytesIO(build_xlsx(DAILY_ROWS))
        upload2.name = "export.xlsx"
        again = services.import_consumption_file(household, user, meter=meter, uploaded_file=upload2)
        assert again.created_count == 0
        assert again.skipped_count == 3

    def test_daily_points_visible_in_day_view_but_not_hour_view(self):
        household, user = HouseholdFactory(), UserFactory()
        meter = ElectricityMeterFactory(household=household, timezone="Europe/Paris")
        upload = io.BytesIO(build_xlsx(DAILY_ROWS))
        upload.name = "export.xlsx"
        services.import_consumption_file(household, user, meter=meter, uploaded_file=upload)

        day = services.consumption_summary(
            household, meter, granularity="day",
            date_from=dt.date(2026, 6, 5), date_to=dt.date(2026, 6, 7),
        )
        assert day["total_wh"] == 15064 + 12485 + 12406
        assert day["estimated_wh"] == 0

        hour = services.consumption_summary(
            household, meter, granularity="hour",
            date_from=dt.date(2026, 6, 5), date_to=dt.date(2026, 6, 7),
        )
        assert hour["total_wh"] == 0

    def test_preview_detects_xlsx(self):
        preview = services.preview_consumption_file(build_xlsx(DAILY_ROWS))
        assert preview["detected_provider"] == "enedis_xlsx"
        assert preview["sample_lines"][0].startswith("05/06/2026")
