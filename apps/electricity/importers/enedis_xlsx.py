"""Enedis XLSX energy export ("Export des données de l'énergie" from the
client portal, e.g. ``…_Export_energie_Consommation_<dates>.xlsx``).

Shape verified against a real household export: a cover sheet, then a data
sheet named ``Export Consommation Quotidienne`` (or another ``Export …``
variant) holding a header row ``Date | Valeur (en kWh)`` followed by one row
per period. Daily rows carry the day's energy in kWh (interval = 24 h,
``ts_start`` = local midnight); if the export ever carries finer timestamps,
the step is inferred from consecutive rows. Dates come as ``DD/MM/YYYY``
strings or real datetime cells depending on the export.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

from .base import BaseImporter, ImporterError, ImporterFormatError, NormalizedPoint
from .registry import register

XLSX_MAGIC = b"PK\x03\x04"
DAY_MINUTES = 24 * 60


def _load_workbook(raw: bytes):
    import openpyxl

    try:
        return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # zip/format errors from openpyxl are varied
        raise ImporterFormatError(f"unreadable XLSX file: {exc}")


def _find_data_sheet(workbook):
    for name in workbook.sheetnames:
        if name.strip().lower().startswith("export"):
            return workbook[name]
    raise ImporterFormatError("no 'Export …' data sheet found — not an Enedis energy export")


def _parse_date_cell(value, line_no: int) -> date | datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
        except ValueError:
            continue
        return parsed if fmt != "%d/%m/%Y" else parsed.date()
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        raise ImporterError(f"row {line_no}: unreadable date {text!r}")


def _parse_value_cell(value, line_no: int) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.upper() == "NA":
        return None  # measurement gap
    try:
        return float(text.replace(",", "."))
    except ValueError:
        raise ImporterError(f"row {line_no}: unreadable value {text!r}")


class EnedisXlsxImporter(BaseImporter):
    key = "enedis_xlsx"
    label = "Enedis — export énergie (XLSX)"

    def detect(self, raw: bytes) -> bool:
        if not raw.startswith(XLSX_MAGIC):
            return False
        try:
            sheet = _find_data_sheet(_load_workbook(raw))
            self._data_rows(sheet)  # requires the Date/Valeur header too
            return True
        except ImporterError:
            return False

    def parse(self, raw: bytes, *, tz: ZoneInfo, options: dict | None = None) -> list[NormalizedPoint]:
        sheet = _find_data_sheet(_load_workbook(raw))
        rows = self._data_rows(sheet)

        # first pass: timestamps + kWh values, gaps skipped
        parsed: list[tuple[datetime, float]] = []
        for line_no, date_cell, value_cell in rows:
            value = _parse_value_cell(value_cell, line_no)
            if value is None:
                continue
            if value < 0:
                raise ImporterError(f"row {line_no}: negative value {value!r}")
            moment = _parse_date_cell(date_cell, line_no)
            if not isinstance(moment, datetime):
                moment = datetime.combine(moment, datetime.min.time())
            if moment.tzinfo is None:
                moment = moment.replace(tzinfo=tz)
            parsed.append((moment, value))

        if not parsed:
            return []

        step = self._infer_step_minutes([moment for moment, _ in parsed])
        return [
            NormalizedPoint(
                ts_start=moment,
                interval_minutes=step,
                energy_wh=int(round(value * 1000)),
                register="base",
            )
            for moment, value in parsed
        ]

    def sample_lines(self, raw: bytes) -> list[str]:
        try:
            sheet = _find_data_sheet(_load_workbook(raw))
        except ImporterError:
            return []
        return [f"{d} ; {v}" for _, d, v in self._data_rows(sheet)[:10]]

    @staticmethod
    def _data_rows(sheet) -> list[tuple[int, object, object]]:
        """Locate the ``Date | Valeur`` header, return (row_no, date, value) rows."""
        rows = []
        header_cols: tuple[int, int] | None = None
        for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if header_cols is None:
                lowered = [c.lower() for c in cells]
                if "date" in lowered and any(c.startswith("valeur") for c in lowered):
                    date_idx = lowered.index("date")
                    value_idx = next(i for i, c in enumerate(lowered) if c.startswith("valeur"))
                    header_cols = (date_idx, value_idx)
                continue
            date_cell, value_cell = row[header_cols[0]], row[header_cols[1]]
            if date_cell is None and value_cell is None:
                continue
            if date_cell is None:
                continue
            rows.append((row_no, date_cell, value_cell))
        if header_cols is None:
            raise ImporterFormatError(
                "no 'Date / Valeur' header found — not an Enedis energy export"
            )
        return rows

    @staticmethod
    def _infer_step_minutes(timestamps: list[datetime]) -> int:
        if len(timestamps) < 2:
            return DAY_MINUTES
        gaps: dict[int, int] = {}
        for a, b in zip(timestamps, timestamps[1:]):
            minutes = int((b - a).total_seconds() // 60)
            if minutes > 0:
                gaps[minutes] = gaps.get(minutes, 0) + 1
        if not gaps:
            return DAY_MINUTES
        # cap at one day: daily exports separated by DST days must not drift
        return min(max(gaps, key=gaps.get), DAY_MINUTES)


register(EnedisXlsxImporter())
